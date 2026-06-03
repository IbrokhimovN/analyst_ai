"""Analitikani Excel (.xlsx) ga eksport qilish — openpyxl bilan."""
import io

from apps.analytics.services import AnalyticsService


def _autosize(ws):
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None),
                    default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 50)


def build_analytics_workbook(source=None, period=None) -> bytes:
    """Umumiy, menejerlar va voronka varaqlari bilan Excel kitobini qaytaradi."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    svc = AnalyticsService()
    wb = Workbook()
    bold = Font(bold=True)

    # 1-varaq: Umumiy
    ws = wb.active
    ws.title = 'Umumiy'
    summary = svc.get_summary(source=source, period=period)
    ws.append(['Ko\'rsatkich', 'Qiymat'])
    for c in ws[1]:
        c.font = bold
    labels = {
        'total_leads': 'Jami lidlar', 'new_leads': 'Yangi lidlar',
        'won_count': 'Yutilgan', 'lost_count': 'Yutqazilgan',
        'conversion_rate': 'Konversiya %', 'total_revenue': 'Umumiy tushum',
        'avg_deal_size': "O'rtacha chek",
    }
    for key, label in labels.items():
        ws.append([label, summary.get(key, 0)])
    _autosize(ws)

    # 2-varaq: Menejerlar
    ws2 = wb.create_sheet('Menejerlar')
    headers = ['Menejer', 'Lidlar', 'Yutilgan', 'Yutqazilgan',
               'Konversiya %', 'Tushum']
    ws2.append(headers)
    for c in ws2[1]:
        c.font = bold
    for m in svc.get_by_manager(source=source, period=period):
        ws2.append([m['manager_name'], m['total_leads'], m['won'],
                    m['lost'], m['conversion_rate'], m['revenue']])
    _autosize(ws2)

    # 3-varaq: Voronka
    ws3 = wb.create_sheet('Voronka')
    ws3.append(['Bosqich', 'Soni', 'Foiz %'])
    for c in ws3[1]:
        c.font = bold
    for s in svc.get_sales_funnel(source=source, period=period):
        ws3.append([s['name'], s['count'], s['pct']])
    _autosize(ws3)

    # 4-varaq: Pipeline kesimi
    ws4 = wb.create_sheet('Pipelinelar')
    ws4.append(['Pipeline', 'Lidlar', 'Yutilgan', 'Yutqazilgan', 'Konversiya %'])
    for c in ws4[1]:
        c.font = bold
    for p in svc.get_pipeline_breakdown(source=source, period=period):
        ws4.append([p['pipeline'], p['leads'], p['won'], p['lost'],
                    p['conversion_rate']])
    _autosize(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
